#!/usr/bin/env python3
"""
VMware2KVM 設計書 (design_document.xlsx) 生成スクリプト
ソースコード解析結果に基づき、5シート構成の設計書をExcelファイルとして出力する。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# スタイル定義
# ============================================================
HEADER_FONT = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
SUBHEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Yu Gothic", size=10)
TITLE_FONT = Font(name="Yu Gothic", size=14, bold=True, color="2C3E50")
MERMAID_FONT = Font(name="Consolas", size=9)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def style_subheader_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, col_count):
    alt_fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER
            if (r - start_row) % 2 == 1:
                cell.fill = alt_fill


def auto_width(ws, col_count, max_width=60, min_width=12):
    for c in range(1, col_count + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    longest = max(len(l) for l in lines)
                    max_len = max(max_len, min(longest + 2, max_width))
        ws.column_dimensions[get_column_letter(c)].width = max_len


# ============================================================
# Sheet 1: 機能一覧表
# ============================================================
def create_feature_list(wb):
    ws = wb.create_sheet("1_機能一覧表")
    ws.sheet_properties.tabColor = "2C3E50"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "VMware2KVM 機能一覧表（基本設計）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["機能ID", "機能名", "機能概要", "対象レイヤー", "対象ユーザー", "関連ファイル"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    features = [
        ["F-001", "VMXファイル解析",
         "VMware .vmx 設定ファイルを読み込み、CPU/メモリ/ディスク/ネットワーク/ファームウェア等のVM構成情報を抽出する。複数エンコーディング(UTF-8, Shift_JIS, CP932, Latin-1)に対応。",
         "バックエンド\n(Parser)", "全ユーザー", "src/vmx_parser.py"],
        ["F-002", "OVF/OVAファイル解析",
         "OVF (XML) を解析しVM構成を抽出。OVA (.tar) は自動展開して内部の .ovf を解析する。CIM ResourceType コードに基づくハードウェアリソース判定を実装。",
         "バックエンド\n(Parser)", "全ユーザー", "src/ovf_parser.py"],
        ["F-003", "KVM libvirt XML生成",
         "VMConfigデータモデルからlibvirt準拠のドメイン定義XMLを生成する。CPU トポロジー、OVMF/UEFI、Hyper-V enlightenments、VNC、qemu-guest-agent等を含む。",
         "バックエンド\n(Generator)", "KVM管理者", "src/kvm_generator.py"],
        ["F-004", "KVMインポートスクリプト生成",
         "qemu-img 変換コマンド + virsh define/start を含むシェルスクリプトを生成。Windows VMの場合はvirtioドライバ導入手順も記載。",
         "バックエンド\n(Generator)", "KVM管理者", "src/kvm_generator.py"],
        ["F-005", "Nutanix acliスクリプト生成",
         "Nutanix CVM上で実行可能なacliコマンドスクリプトを生成。Image Service登録→アップロード完了待機→VM作成→ディスククローン→NICアタッチ→検証の完全フローを含む。",
         "バックエンド\n(Generator)", "Nutanix管理者", "src/nutanix_generator.py"],
        ["F-006", "Nutanix REST APIスクリプト生成",
         "Prism Central v3 REST API を用いた curl ベースのインポートスクリプトを生成。Image UUID取得、COMPLETE状態待機、VM作成を自動化。",
         "バックエンド\n(Generator)", "Nutanix管理者", "src/nutanix_generator.py"],
        ["F-007", "Nutanix VM Spec JSON生成",
         "Prism v3 API 用の VM 仕様 JSON を生成。disk_list, nic_list, boot_config, categories等を含む。",
         "バックエンド\n(Generator)", "Nutanix管理者", "src/nutanix_generator.py"],
        ["F-008", "一括変換オーケストレーション",
         "入力フォルダ内の全VMファイルをスキャンし、パーサー→ジェネレーターを順次実行。停止要求、進捗コールバック、ログコールバックに対応。",
         "バックエンド\n(Orchestrator)", "全ユーザー", "src/converter.py"],
        ["F-009", "ディスクイメージ変換 (vmdk→qcow2)",
         "qemu-img コマンドを呼び出し、VMDKをqcow2形式に変換する。圧縮オプション(-c)対応。qemu-img未導入時はコマンド文字列のみ生成。",
         "バックエンド\n(Orchestrator)", "全ユーザー", "src/converter.py"],
        ["F-010", "Windows BSOD回避機能",
         "Windows ゲストOS検出時、ディスクバスをSATA(EFI)/IDE(BIOS)にフォールバックしBSODを防止。virtio-win.isoをCD-ROMに自動マウント。Hyper-V Enlightenments、localtime clock を自動設定。",
         "バックエンド\n(Generator)", "KVM管理者", "src/kvm_generator.py"],
        ["F-011", "ネットワークマッピング",
         "外部JSONファイル(network_map.json)を読み込み、VMwareポートグループ名をKVMブリッジ名/Nutanixサブネットに変換。MACアドレスはXML/JSON/acliすべてに引き継ぐ。",
         "バックエンド\n(Cross-cutting)", "インフラ管理者", "network_map.json\nsrc/kvm_generator.py\nsrc/nutanix_generator.py"],
        ["F-012", "Pre-flight Check（事前チェック）",
         "変換前にVMDK合計サイズと出力先ディスク空き容量を比較。容量不足の場合は警告を出して処理を中断する。GUI/CLI両対応。",
         "バックエンド\n(Orchestrator)", "全ユーザー", "src/converter.py"],
        ["F-013", "移行レポート出力",
         "一括変換結果をmigration_report.csv(CSV)とmigration.log(テキスト)として出力。VM名、状態、CPU、メモリ、NIC/MACアドレス、所要時間、エラー内容を記録。",
         "バックエンド\n(Orchestrator)", "全ユーザー", "src/converter.py"],
        ["F-014", "GUIモード",
         "tkinter ベースのデスクトップGUIを提供。入出力フォルダ選択、変換先/オプション設定、プログレスバー、リアルタイムログ表示、停止機能を実装。MVCアーキテクチャ。",
         "フロントエンド\n(GUI)", "全ユーザー", "src/gui_view.py\nsrc/gui_controller.py\nsrc/gui_model.py"],
        ["F-015", "CLIモード",
         "コマンドライン引数によるバッチ実行モードを提供。argparseによる13種のオプションをサポート。スクリプトや自動化パイプラインに組み込み可能。",
         "フロントエンド\n(CLI)", "自動化担当者", "main.py"],
        ["F-016", "多言語対応 (i18n)",
         "日本語/英語の2言語をサポート。翻訳キーは64種。GUIのリアルタイム言語切替に対応。",
         "フロントエンド\n(Cross-cutting)", "全ユーザー", "src/i18n.py"],
        ["F-017", "ゲストOS自動判定",
         "VMware guestOS コード(27種)から OS ファミリー(windows/linux)と libvirt os-variant を自動判定。Windows検出時はBSOD回避・Hyper-V最適化を自動適用。",
         "バックエンド\n(Parser)", "全ユーザー", "src/vmx_parser.py\nGUEST_OS_MAP"],
    ]

    for r, row_data in enumerate(features, 4):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 4, 3 + len(features), len(headers))
    auto_width(ws, len(headers))
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["F"].width = 30
    ws.freeze_panes = "A4"


# ============================================================
# Sheet 2: API仕様書 (CLI引数 + 内部公開メソッド)
# ============================================================
def create_api_spec(wb):
    ws = wb.create_sheet("2_API仕様書")
    ws.sheet_properties.tabColor = "3498DB"

    ws.merge_cells("A1:H1")
    ws["A1"] = "VMware2KVM API仕様書（詳細設計）— CLI インターフェース & 内部公開メソッド"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    # --- CLI ---
    ws.merge_cells("A3:H3")
    ws["A3"] = "CLIインターフェース (main.py)"
    ws["A3"].font = Font(name="Yu Gothic", size=12, bold=True, color="2C3E50")

    cli_headers = ["オプション", "短縮形", "型", "デフォルト", "必須", "選択肢", "説明", "実装箇所"]
    for i, h in enumerate(cli_headers, 1):
        ws.cell(row=5, column=i, value=h)
    style_header_row(ws, 5, len(cli_headers))

    cli_args = [
        ["--input", "-i", "str (path)", "-", "CLI時必須", "-", "入力フォルダ（VMwareファイル配置先）", "main.py → ConversionOptions"],
        ["--output", "-o", "str (path)", "-", "CLI時必須", "-", "出力フォルダ（変換結果の出力先）", "main.py → ConversionOptions"],
        ["--target", "-t", "str", "kvm", "No", "kvm / nutanix / both", "変換先プラットフォームの選択", "ConversionOptions.target"],
        ["--ext", "-", "str (CSV)", ".vmx,.ovf,.ova", "No", "-", "スキャン対象のファイル拡張子（カンマ区切り）", "ConversionOptions.extensions"],
        ["--no-disk", "-", "flag", "False", "No", "-", "ディスクイメージ変換をスキップ（設定ファイルのみ生成）", "ConversionOptions.convert_disk"],
        ["--no-compress", "-", "flag", "False", "No", "-", "qcow2 圧縮を無効化", "ConversionOptions.compress_qcow2"],
        ["--container", "-", "str", "default-container", "No", "-", "Nutanix コンテナ名", "ConversionOptions.nutanix_container"],
        ["--network-map", "-", "str (path)", "auto", "No", "-", "network_map.json のパス。未指定時はBASE_DIR配下を自動検索", "ConversionOptions.network_map_path"],
        ["--no-win-fallback", "-", "flag", "False", "No", "-", "Windows VM の SATA/IDE フォールバックを無効化（上級者向け）", "ConversionOptions.windows_bus_fallback"],
        ["--virtio-iso", "-", "str (path)", "auto", "No", "-", "virtio-win.iso のファイルパス", "ConversionOptions.virtio_win_iso"],
        ["--lang", "-", "str", "ja", "No", "ja / en", "UIおよびログの表示言語", "I18n.lang"],
        ["--gui", "-", "flag", "False", "No", "-", "CLIオプション指定時でもGUIモードを強制起動", "main.py run_gui()"],
    ]

    for r, row_data in enumerate(cli_args, 6):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 6, 5 + len(cli_args), len(cli_headers))

    # --- Internal methods ---
    method_start = 6 + len(cli_args) + 2
    ws.merge_cells(f"A{method_start}:H{method_start}")
    ws[f"A{method_start}"] = "内部公開メソッド一覧"
    ws[f"A{method_start}"].font = Font(name="Yu Gothic", size=12, bold=True, color="2C3E50")

    meth_headers = ["クラス", "メソッド名", "引数", "戻り値", "処理概要", "例外/エラー", "呼び出し元", "所属ファイル"]
    hr = method_start + 2
    for i, h in enumerate(meth_headers, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(meth_headers))

    methods = [
        ["VmxParser", "parse(vmx_path)", "vmx_path: str", "VMConfig", ".vmxファイルを読み込みVMConfigを返す", "UnicodeDecodeError → フォールバック", "Converter._convert_single", "vmx_parser.py"],
        ["OvfParser", "parse(path)", "path: str", "(VMConfig, str)", "OVF/OVAを解析。OVAは自動展開。戻り値タプル (config, working_dir)", "ValueError: OVF not found in OVA", "Converter._convert_single", "ovf_parser.py"],
        ["KvmGenerator", "generate(config, ...)", "config: VMConfig\noutput_dir: str\ncompress, thin, network_map,\nwindows_bus_fallback, virtio_win_iso", "dict", "libvirt XML + import.sh + disk_commands を生成し結果dictを返す", "-", "Converter._convert_single", "kvm_generator.py"],
        ["NutanixGenerator", "generate(config, ...)", "config: VMConfig\noutput_dir: str\ncompress, container_name,\nnetwork_map", "dict", "acli.sh + api.sh + spec.json + disk_commands を生成し結果dictを返す", "-", "Converter._convert_single", "nutanix_generator.py"],
        ["Converter", "scan_vms(input_dir, ext)", "input_dir: str\nextensions: List[str]", "List[str]", "入力ディレクトリを再帰走査し対象ファイルパスのリストを返す", "-", "convert_all / main.py", "converter.py"],
        ["Converter", "preflight_check(output_dir, vm_files, log_fn)", "output_dir: str\nvm_files: List[str]\nlog_fn: Callable", "bool", "VMDK合計サイズと出力先空き容量を比較。不足ならFalse", "OSError → True (チェック不可時は続行)", "main.py / gui_controller.py", "converter.py"],
        ["Converter", "estimate_total_vmdk_size(vm_files)", "vm_files: List[str]", "int (bytes)", "VMファイルを解析し全VMDKの合計バイト数を算出", "Exception → skip", "preflight_check", "converter.py"],
        ["Converter", "convert_all(input_dir, output_dir, options, ...)", "input_dir, output_dir: str\noptions: ConversionOptions\nlog_fn, progress_fn: Callable", "List[ConversionResult]", "全VMを一括変換。ネットワークマップ読み込み→ループ変換→レポート出力", "-", "main.py / gui_controller.py", "converter.py"],
        ["Converter", "load_network_map(path)", "path: str", "Dict", "JSONファイルを読み込みネットワークマッピング辞書を返す", "-", "convert_all", "converter.py"],
        ["Converter", "stop()", "-", "None", "変換ループの停止フラグをセット", "-", "gui_controller.py", "converter.py"],
        ["Converter", "_write_report(output_dir, results)", "output_dir: str\nresults: List[ConversionResult]", "None", "migration_report.csv と migration.log を出力", "-", "convert_all", "converter.py"],
        ["I18n", "t(key, **kwargs)", "key: str, **kwargs", "str", "翻訳キーを現在言語で解決。{placeholder}をkwargsで置換", "-", "View / Controller / main.py", "i18n.py"],
        ["I18n", "set_lang(lang)", "lang: str", "None", "表示言語を切り替え", "-", "gui_controller.py", "i18n.py"],
    ]

    for r, row_data in enumerate(methods, hr + 1):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, hr + 1, hr + len(methods), len(meth_headers))
    auto_width(ws, len(meth_headers))
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A6"


# ============================================================
# Sheet 3: テーブル定義書 (データモデル)
# ============================================================
def create_table_definitions(wb):
    ws = wb.create_sheet("3_テーブル定義書")
    ws.sheet_properties.tabColor = "27AE60"

    ws.merge_cells("A1:G1")
    ws["A1"] = "VMware2KVM データモデル定義書（詳細設計）— dataclass / 設定ファイル構造"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    tables = [
        ("VMConfig", "src/vmx_parser.py", "VM構成の統一データモデル。全パーサーの出力・全ジェネレーターの入力として使用される中核モデル。", [
            ["name", "str", "必須", "-", "VM内部名（サニタイズ済み、ファイル名安全）"],
            ["display_name", "str", "必須", "-", "VM表示名（元の.vmxのdisplayName）"],
            ["guest_os", "str", "必須", "-", "VMwareゲストOSコード (例: rhel8-64, windows2019srvnext-64)"],
            ["num_cpus", "int", "必須", "-", "仮想CPU数"],
            ["cores_per_socket", "int", "必須", "-", "ソケットあたりのコア数"],
            ["memory_mb", "int", "必須", "-", "メモリサイズ (MB)"],
            ["disks", "List[DiskInfo]", "-", "[]", "接続ディスク一覧"],
            ["networks", "List[NetworkInfo]", "-", "[]", "接続ネットワーク一覧"],
            ["firmware", "str", "-", '"bios"', "ファームウェア種別 (bios / efi)"],
            ["source_file", "str", "-", '""', "解析元ファイルのフルパス"],
            ["hardware_version", "int", "-", "0", "VMwareハードウェアバージョン"],
            ["annotations", "str", "-", '""', "VMの注釈/メモ"],
            ["extra", "Dict[str, str]", "-", "{}", "未分類のvmxプロパティ"],
        ]),
        ("DiskInfo", "src/vmx_parser.py", "VMに接続された仮想ディスクの情報。", [
            ["filename", "str", "必須", "-", "VMDKファイルパス（絶対パスに解決済み）"],
            ["adapter_type", "str", "必須", "-", "コントローラ種別 (scsi / ide / sata / nvme)"],
            ["controller_id", "int", "必須", "-", "コントローラID (0-3)"],
            ["unit_id", "int", "必須", "-", "ユニットID (0-15)"],
            ["size_kb", "int", "-", "0", "ディスクサイズ (KB)"],
            ["mode", "str", "-", '"persistent"', "ディスクモード"],
            ["thin", "bool", "-", "False", "シンプロビジョニングフラグ"],
        ]),
        ("NetworkInfo", "src/vmx_parser.py", "VMに接続されたネットワークアダプタの情報。", [
            ["adapter_type", "str", "必須", "-", "アダプタ種別 (e1000 / vmxnet3 / e1000e 等)"],
            ["network_name", "str", "必須", "-", "VMwareポートグループ名 (例: VM Network)"],
            ["mac_address", "str", "-", '""', "MACアドレス（移行先に引き継ぎ）"],
            ["connected", "bool", "-", "True", "接続状態"],
        ]),
        ("ConversionOptions", "src/converter.py", "変換実行時のオプション設定。CLIArgs / GUIウィジェットから生成される。", [
            ["target", "str", "-", '"kvm"', "変換先: kvm / nutanix / both"],
            ["convert_disk", "bool", "-", "True", "ディスクイメージ変換を実行するか"],
            ["generate_config", "bool", "-", "True", "VM設定ファイルを生成するか"],
            ["generate_script", "bool", "-", "True", "インポートスクリプトを生成するか"],
            ["compress_qcow2", "bool", "-", "True", "qcow2圧縮を有効にするか"],
            ["thin_provision", "bool", "-", "True", "シンプロビジョニング"],
            ["nutanix_container", "str", "-", '"default-container"', "Nutanixコンテナ名"],
            ["extensions", "List[str]", "-", '[".vmx",".ovf",".ova"]', "スキャン対象の拡張子リスト"],
            ["network_map_path", "str", "-", '""', "network_map.json のパス"],
            ["windows_bus_fallback", "bool", "-", "True", "Windows BSOD回避フォールバック"],
            ["virtio_win_iso", "str", "-", '""', "virtio-win.iso のパス"],
        ]),
        ("ConversionResult", "src/converter.py", "1VM分の変換結果。convert_allの戻り値リスト要素。", [
            ["vm_name", "str", "必須", "-", "VM名"],
            ["success", "bool", "必須", "-", "変換成否"],
            ["config", "Optional[VMConfig]", "-", "None", "解析済みVM構成"],
            ["kvm_result", "Optional[dict]", "-", "None", "KVM生成結果 (xml_path, disk_commands等)"],
            ["nutanix_result", "Optional[dict]", "-", "None", "Nutanix生成結果 (acli_script, api_script等)"],
            ["error", "str", "-", '""', "エラーメッセージ"],
            ["disk_converted", "bool", "-", "False", "ディスク変換成否"],
            ["elapsed_sec", "float", "-", "0.0", "処理所要時間（秒）"],
        ]),
        ("ConversionState", "src/gui_model.py", "GUI変換状態の管理モデル。Controllerが更新しViewが参照する。", [
            ["running", "bool", "-", "False", "変換実行中フラグ"],
            ["total", "int", "-", "0", "変換対象VM総数"],
            ["current", "int", "-", "0", "現在処理中のインデックス"],
            ["success", "int", "-", "0", "成功数"],
            ["errors", "List[str]", "-", "[]", "エラーメッセージリスト"],
        ]),
        ("network_map.json", "network_map.json", "ネットワークマッピング外部設定ファイル。VMwareポートグループ名をKVM/Nutanix環境のネットワークに変換する。", [
            ['kvm.{ネットワーク名}.type', "str", "-", '-', "KVM接続タイプ (bridge / network)"],
            ['kvm.{ネットワーク名}.name', "str", "-", '-', "KVMブリッジ名またはlibvirtネットワーク名"],
            ['nutanix.{ネットワーク名}.subnet_name', "str", "-", '-', "Nutanixサブネット名"],
            ['nutanix.{ネットワーク名}.subnet_uuid', "str", "-", '-', "NutanixサブネットUUID"],
            ['nutanix.{ネットワーク名}.vlan_id', "int", "-", '0', "VLAN ID"],
        ]),
        ("GUEST_OS_MAP", "src/vmx_parser.py", "VMware guestOSコードから(OSファミリー, libvirt os-variant)へのマッピング定数。27エントリ。", [
            ["キー", "値", "-", "-", "マッピング例"],
            ["windows9-64", '("windows", "win10")', "-", "-", "Windows 10 64bit"],
            ["windows2019srvnext-64", '("windows", "win2k22")', "-", "-", "Windows Server 2022"],
            ["rhel8-64", '("linux", "rhel8.0")', "-", "-", "Red Hat Enterprise Linux 8"],
            ["centos8-64", '("linux", "centos8")', "-", "-", "CentOS 8"],
            ["ubuntu-64", '("linux", "ubuntu20.04")', "-", "-", "Ubuntu 64bit"],
            ["other", '("linux", "generic")', "-", "-", "不明なOS → generic"],
        ]),
    ]

    row = 3
    for tbl_name, tbl_file, tbl_desc, fields in tables:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = f"{tbl_name}  ({tbl_file})"
        ws[f"A{row}"].font = Font(name="Yu Gothic", size=11, bold=True, color="2C3E50")
        row += 1

        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = tbl_desc
        ws[f"A{row}"].font = Font(name="Yu Gothic", size=9, italic=True)
        row += 1

        fld_headers = ["フィールド名", "データ型", "制約", "デフォルト値", "論理名 / 説明"]
        for i, h in enumerate(fld_headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_subheader_row(ws, row, len(fld_headers))
        row += 1

        for fld in fields:
            for c, val in enumerate(fld, 1):
                ws.cell(row=row, column=c, value=val)
            row += 1

        style_body(ws, row - len(fields), row - 1, len(fld_headers))
        row += 1

    auto_width(ws, 7)
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A3"


# ============================================================
# Sheet 4: エラー・ログ定義書
# ============================================================
def create_error_log_definitions(wb):
    ws = wb.create_sheet("4_エラー・ログ定義書")
    ws.sheet_properties.tabColor = "E74C3C"

    ws.merge_cells("A1:G1")
    ws["A1"] = "VMware2KVM エラー・ログ定義書（詳細設計）"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["i18nキー", "カテゴリ", "ログレベル", "日本語メッセージ", "英語メッセージ", "出力タイミング", "発生箇所"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    entries = [
        # エラー系
        ["error.no_input", "入力エラー", "ERROR", "入力フォルダを選択してください", "Please select an input folder", "変換実行ボタン押下時", "gui_controller.py"],
        ["error.no_output", "入力エラー", "ERROR", "出力フォルダを選択してください", "Please select an output folder", "変換実行ボタン押下時", "gui_controller.py"],
        ["error.input_not_found", "入力エラー", "ERROR", "入力フォルダが存在しません", "Input folder does not exist", "変換実行ボタン押下時", "gui_controller.py"],
        ["error.no_vms", "入力エラー", "INFO", "VMwareファイルが見つかりません", "No VMware files found", "スキャン後VM数が0の場合", "gui_controller.py / main.py"],
        ["error.qemu_not_found", "外部依存", "WARN", "qemu-img が見つかりません。コマンド生成のみ行います。", "qemu-img not found. Generating commands only.", "ディスク変換実行時にPATHにqemu-imgがない", "converter.py"],
        ["error.disk_space", "容量エラー", "ERROR", "出力先のディスク容量が不足しています ({vmdk_size} GB 必要 / {free_size} GB 空き)", "Insufficient disk space ({vmdk_size} GB needed / {free_size} GB free)", "Pre-flightチェック失敗時（GUI）", "gui_controller.py"],
        # ログ系
        ["log.start", "変換開始", "INFO", "=== VMware → {target} 変換開始 ===", "=== VMware → {target} conversion started ===", "convert_all開始時", "gui_controller.py / main.py"],
        ["log.found", "スキャン", "INFO", "{count} 個のVMを検出", "{count} VM(s) found", "scan_vms完了後", "gui_controller.py / main.py"],
        ["log.vm_start", "VM変換", "INFO", "--- VM変換開始: {name} ---", "--- Converting VM: {name} ---", "各VM変換開始時", "converter.py"],
        ["log.parsing_vmx", "解析", "INFO", "VMX解析中: {file}", "Parsing VMX: {file}", ".vmx解析開始時", "converter.py"],
        ["log.parsing_ovf", "解析", "INFO", "OVF解析中: {file}", "Parsing OVF: {file}", ".ovf/.ova解析開始時", "converter.py"],
        ["log.generating_xml", "生成", "INFO", "libvirt XML生成中: {name}", "Generating libvirt XML: {name}", "KVM XML生成開始時", "converter.py"],
        ["log.generating_nutanix", "生成", "INFO", "Nutanixインポートスクリプト生成中: {name}", "Generating Nutanix import script: {name}", "Nutanixスクリプト生成開始時", "converter.py"],
        ["log.converting_disk", "ディスク", "INFO", "ディスク変換中: {file}", "Converting disk: {file}", "qemu-img実行中", "converter.py"],
        ["log.disk_done", "ディスク", "INFO", "ディスク変換完了: {file}", "Disk conversion done: {file}", "qemu-img正常完了時", "converter.py"],
        ["log.skip_disk", "ディスク", "WARN", "ディスク変換スキップ (手動でqemu-imgを実行してください)", "Disk conversion skipped (run qemu-img manually)", "qemu-img未導入時", "converter.py"],
        ["log.cmd_hint", "情報", "DEBUG", "実行コマンド: {cmd}", "Command: {cmd}", "生成コマンドの表示", "converter.py"],
        ["log.vm_done", "VM変換", "INFO", "VM変換完了: {name}", "VM conversion done: {name}", "各VM変換完了時", "converter.py"],
        ["log.complete", "変換完了", "INFO", "=== 変換完了: {success}/{total} 成功 ===", "=== Conversion complete: {success}/{total} succeeded ===", "convert_all完了時", "gui_controller.py"],
        ["log.error", "エラー", "ERROR", "[エラー] {message}", "[ERROR] {message}", "例外・失敗発生時", "converter.py / gui_controller.py"],
        ["log.preflight_disk", "事前チェック", "INFO", "事前チェック: VMDK合計 {vmdk_size} GB / 出力先空き {free_size} GB", "Pre-flight: VMDK total {vmdk_size} GB / Output free {free_size} GB", "Pre-flightチェック実行時", "converter.py"],
        ["log.preflight_fail", "事前チェック", "ERROR", "容量不足! VMDK合計 {vmdk_size} GB > 空き {free_size} GB。処理を中断します。", "Insufficient space! VMDK total {vmdk_size} GB > Free {free_size} GB. Aborting.", "Pre-flightチェック失敗時", "converter.py"],
        # ステータス系
        ["status.ready", "ステータス", "INFO", "準備完了", "Ready", "GUI初期表示時", "gui_view.py"],
        ["status.converting", "ステータス", "INFO", "変換中: {file} ({current}/{total})", "Converting: {file} ({current}/{total})", "変換ループ中", "gui_controller.py"],
        ["status.done", "ステータス", "INFO", "完了: {success}/{total} 成功", "Done: {success}/{total} succeeded", "変換完了時", "gui_controller.py"],
        ["status.error", "ステータス", "ERROR", "エラー: {message}", "Error: {message}", "予期しない例外発生時", "gui_controller.py"],
        ["status.stopped", "ステータス", "WARN", "停止されました", "Stopped", "ユーザーが停止ボタン押下時", "gui_controller.py"],
    ]

    for r, row_data in enumerate(entries, 4):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 4, 3 + len(entries), len(headers))
    auto_width(ws, len(headers))
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 35
    ws.freeze_panes = "A4"


# ============================================================
# Sheet 5: アーキテクチャ図解 (Mermaid)
# ============================================================
def create_architecture_diagrams(wb):
    ws = wb.create_sheet("5_アーキテクチャ図解")
    ws.sheet_properties.tabColor = "8E44AD"

    ws.merge_cells("A1:B1")
    ws["A1"] = "VMware2KVM アーキテクチャ図解（Mermaid記法）"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    diagrams = [
        ("システム構成図（コンポーネント図）", """graph TB
    subgraph "VMware2KVM"
        subgraph "Frontend"
            GUI["GUI (tkinter)<br/>gui_view.py"]
            CLI["CLI (argparse)<br/>main.py"]
            I18N["I18n<br/>i18n.py"]
        end
        subgraph "Controller"
            CTRL["ConversionController<br/>gui_controller.py"]
        end
        subgraph "Backend - Parsers"
            VMX["VmxParser<br/>vmx_parser.py"]
            OVF["OvfParser<br/>ovf_parser.py"]
        end
        subgraph "Backend - Generators"
            KVM["KvmGenerator<br/>kvm_generator.py"]
            NTX["NutanixGenerator<br/>nutanix_generator.py"]
        end
        subgraph "Backend - Orchestrator"
            CONV["Converter<br/>converter.py"]
        end
        subgraph "Data Models"
            CFG["VMConfig / DiskInfo / NetworkInfo"]
            OPT["ConversionOptions / ConversionResult"]
        end
        subgraph "Config"
            NMAP["network_map.json"]
        end
    end

    GUI --> CTRL
    CLI --> CONV
    CTRL --> CONV
    CONV --> VMX
    CONV --> OVF
    CONV --> KVM
    CONV --> NTX
    VMX --> CFG
    OVF --> CFG
    KVM --> CFG
    NTX --> CFG
    CONV --> OPT
    CONV --> NMAP
    GUI --> I18N
    CLI --> I18N"""),

        ("変換処理シーケンス図", """sequenceDiagram
    participant U as User
    participant M as main.py / GUI
    participant C as Converter
    participant P as VmxParser / OvfParser
    participant KG as KvmGenerator
    participant NG as NutanixGenerator
    participant FS as FileSystem

    U->>M: 変換実行 (input_dir, output_dir, options)
    M->>C: convert_all(input_dir, output_dir, options)
    C->>C: load_network_map(path)
    C->>C: scan_vms(input_dir, extensions)
    C-->>M: progress_fn(0, total)

    loop 各VMファイル
        C->>C: time.time() → t_start
        C->>P: parse(vm_file)
        P->>FS: read .vmx / .ovf / .ova
        P-->>C: VMConfig

        alt target = kvm or both
            C->>KG: generate(config, output_dir, ...)
            KG->>FS: write {name}.xml
            KG->>FS: write import_{name}.sh
            KG-->>C: kvm_result dict
        end

        alt target = nutanix or both
            C->>NG: generate(config, output_dir, ...)
            NG->>FS: write import_{name}_acli.sh
            NG->>FS: write import_{name}_api.sh
            NG->>FS: write {name}_spec.json
            NG-->>C: nutanix_result dict
        end

        C->>C: elapsed_sec = time.time() - t_start
        C-->>M: progress_fn(i, total)
    end

    C->>FS: write migration_report.csv
    C->>FS: write migration.log
    C-->>M: List[ConversionResult]"""),

        ("データモデルER図", """erDiagram
    VMConfig ||--o{ DiskInfo : "has disks"
    VMConfig ||--o{ NetworkInfo : "has networks"
    VMConfig {
        str name PK
        str display_name
        str guest_os
        int num_cpus
        int cores_per_socket
        int memory_mb
        str firmware
        str source_file
        int hardware_version
        str annotations
    }
    DiskInfo {
        str filename
        str adapter_type
        int controller_id
        int unit_id
        int size_kb
        str mode
        bool thin
    }
    NetworkInfo {
        str adapter_type
        str network_name
        str mac_address
        bool connected
    }
    ConversionOptions {
        str target
        bool convert_disk
        bool generate_config
        bool compress_qcow2
        str nutanix_container
        str network_map_path
        bool windows_bus_fallback
        str virtio_win_iso
    }
    ConversionResult {
        str vm_name
        bool success
        float elapsed_sec
        str error
        bool disk_converted
    }
    ConversionResult ||--o| VMConfig : "references"
    ConversionOptions --o| Converter : "input to"
    Converter ||--o{ ConversionResult : "produces" """),

        ("GUI MVC 構成図", """graph LR
    subgraph Model
        M1["ConversionState<br/>(running, total, current, success, errors)"]
        M2["ConversionOptions"]
        M3["TARGET_OPTIONS / DEFAULT_EXTENSIONS"]
    end
    subgraph View
        V1["VMware2KVMView<br/>tkinter widgets"]
        V2["ScrolledText (log)"]
        V3["Progressbar"]
        V4["Combobox (target/lang)"]
        V5["Checkbuttons (options)"]
    end
    subgraph Controller
        C1["ConversionController<br/>event handling"]
    end
    subgraph Backend
        B1["Converter"]
    end

    V1 -->|bind_command| C1
    V4 -->|<<ComboboxSelected>>| C1
    C1 -->|read vars| V1
    C1 -->|append_log / set_running| V1
    C1 -->|threading.Thread| B1
    B1 -->|root.after callback| V1
    C1 -->|update| M1"""),

        ("ファイル出力マップ", """graph TD
    INPUT["input/<br/>.vmx / .ovf / .ova"] --> CONV["Converter"]

    CONV --> KVM_OUT["KVM Output"]
    CONV --> NTX_OUT["Nutanix Output"]
    CONV --> REPORT["Reports"]

    KVM_OUT --> XML["{name}.xml<br/>libvirt domain XML"]
    KVM_OUT --> KSH["import_{name}.sh<br/>virsh define + start"]

    NTX_OUT --> ACLI["import_{name}_acli.sh<br/>acli commands"]
    NTX_OUT --> API["import_{name}_api.sh<br/>Prism REST API curl"]
    NTX_OUT --> JSON["{name}_spec.json<br/>Prism v3 VM spec"]

    REPORT --> CSV["migration_report.csv"]
    REPORT --> LOG["migration.log"]"""),
    ]

    row = 3
    for title, mermaid_code in diagrams:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Yu Gothic", size=12, bold=True, color="8E44AD")
        row += 1

        ws[f"A{row}"] = "Mermaid Live Editor (https://mermaid.live) にペーストして表示してください"
        ws[f"A{row}"].font = Font(name="Yu Gothic", size=9, italic=True, color="888888")
        row += 1

        ws.merge_cells(f"A{row}:B{row + mermaid_code.count(chr(10)) + 1}")
        ws[f"A{row}"] = mermaid_code
        ws[f"A{row}"].font = MERMAID_FONT
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += mermaid_code.count("\n") + 3

    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 30


# ============================================================
# Main
# ============================================================
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    create_feature_list(wb)
    create_api_spec(wb)
    create_table_definitions(wb)
    create_error_log_definitions(wb)
    create_architecture_diagrams(wb)

    output_path = "design_document.xlsx"
    wb.save(output_path)
    print(f"Design document generated: {output_path}")


if __name__ == "__main__":
    main()
